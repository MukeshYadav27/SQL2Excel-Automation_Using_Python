# report_generator.py

import os
from datetime import datetime

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from config import REPORT_FOLDER


# ---------------------------------------------------------
# Generate Excel Report
# ---------------------------------------------------------

def generate_excel(df, report_name):

    # Create reports folder
    os.makedirs(REPORT_FOLDER, exist_ok=True)

    # File Name
    current_time = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")

    filename = f"{report_name}_{current_time}.xlsx"

    filepath = os.path.join(REPORT_FOLDER, filename)

    # Export dataframe
    df.to_excel(filepath, index=False)

    # Open Workbook
    wb = load_workbook(filepath)

    ws = wb.active

    # Rename Sheet
    ws.title = "Employee Report"

    # -----------------------------------------
    # Header Formatting
    # -----------------------------------------

    header_fill = PatternFill(
        fill_type="solid",
        fgColor="1F4E78"
    )

    header_font = Font(
        bold=True,
        color="FFFFFF"
    )

    for cell in ws[1]:

        cell.fill = header_fill

        cell.font = header_font

        cell.alignment = Alignment(
            horizontal="center",
            vertical="center"
        )

    # -----------------------------------------
    # Auto Adjust Width
    # -----------------------------------------

    for column in ws.columns:

        length = 0

        letter = get_column_letter(column[0].column)

        for cell in column:

            try:

                if len(str(cell.value)) > length:

                    length = len(str(cell.value))

            except:

                pass

        ws.column_dimensions[letter].width = length + 5

    # -----------------------------------------
    # Freeze Header
    # -----------------------------------------

    ws.freeze_panes = "A2"

    wb.save(filepath)

    wb.close()

    return os.path.abspath(filepath)